import openpyxl as xl

class Spreadsheet:

    def __init__(self):
        self.__wb = xl.Workbook()
        self.__sheet = self.__wb.active
        self.__sheet["B2"].value = "Date"
        self.__sheet["C2"].value = "Amount"
        self.__sheet["D2"].value = "Description"


    def insert_data(self, data):
        for row in range(3, 3 + len(data)):
            cell1 = self.__sheet.cell(row, 2)
            cell2 = self.__sheet.cell(row, 3)
            cell3 = self.__sheet.cell(row, 4)
            cell1.value = data[row-3][0]
            cell2.value = data[row-3][1]
            cell3.value = data[row-3][2]


    def save_workbook(self):
        self.__wb.save("temp.xlsx")